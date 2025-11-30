"""
Interest Calculator Tool

Refactored from universal_debt_calculator_cli.py for use as a LangGraph tool.
Supports: simple interest, LPR floating rate, delayed performance, compound, penalty.

LPR data is loaded from the centralized knowledge management system for dynamic updates.
Supports Excel export for detailed calculation process output.
"""

from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP, getcontext
from typing import Dict, List, Tuple, Optional, Union, Any
from enum import Enum
import logging
from pathlib import Path
import yaml

# Excel export support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

logger = logging.getLogger(__name__)

# Set Decimal precision for financial calculations
getcontext().prec = 28
getcontext().rounding = ROUND_HALF_UP


class CalculationType(str, Enum):
    """Supported calculation types."""
    SIMPLE = "simple"
    LPR = "lpr"
    DELAY = "delay"
    COMPOUND = "compound"
    PENALTY = "penalty"
    # 新增计算类型
    SHARE_RATIO = "share_ratio"  # 份额比例计算（银团贷款）
    CONFIRMED = "confirmed"  # 判决确认金额（直接使用，不计算）
    MAX_LIMIT = "max_limit"  # 最高额限额封顶检查


def load_lpr_data_from_yaml() -> List[Tuple[str, float, float]]:
    """
    Load LPR data from the centralized knowledge YAML file.

    Returns:
        List of tuples: (date_str, lpr_1y, lpr_5y)
    """
    # Try to load from knowledge directory
    knowledge_lpr_path = Path(__file__).parent.parent / 'knowledge' / 'calculations' / 'lpr_rates.yaml'

    if knowledge_lpr_path.exists():
        try:
            with open(knowledge_lpr_path, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f)

            rates = []
            for rate_entry in data.get('rates', []):
                date_str = rate_entry.get('date', '')
                lpr_1y = float(rate_entry.get('lpr_1y', 0))
                lpr_5y = float(rate_entry.get('lpr_5y', 0))
                rates.append((date_str, lpr_1y, lpr_5y))

            logger.info(f"Loaded {len(rates)} LPR rates from {knowledge_lpr_path}")
            return rates

        except Exception as e:
            logger.warning(f"Failed to load LPR from YAML: {e}. Using embedded fallback.")

    # Fallback to embedded data if YAML loading fails
    return _get_embedded_lpr_data()


def _get_embedded_lpr_data() -> List[Tuple[str, float, float]]:
    """Fallback embedded LPR data (2019-2025)."""
    return [
        ("2025-07-21", 3.00, 3.50),
        ("2025-06-20", 3.00, 3.50),
        ("2025-05-20", 3.00, 3.50),
        ("2025-04-21", 3.10, 3.60),
        ("2025-03-20", 3.10, 3.60),
        ("2025-02-20", 3.10, 3.60),
        ("2025-01-20", 3.10, 3.60),
        ("2024-12-20", 3.10, 3.60),
        ("2024-11-20", 3.10, 3.60),
        ("2024-10-21", 3.10, 3.60),
        ("2024-09-20", 3.35, 3.85),
        ("2024-08-20", 3.35, 3.85),
        ("2024-07-22", 3.35, 3.85),
        ("2024-06-20", 3.45, 3.95),
        ("2024-05-20", 3.45, 3.95),
        ("2024-04-22", 3.45, 3.95),
        ("2024-03-20", 3.45, 3.95),
        ("2024-02-20", 3.45, 3.95),
        ("2024-01-22", 3.45, 4.20),
        ("2023-12-20", 3.45, 4.20),
        ("2023-11-20", 3.45, 4.20),
        ("2023-10-20", 3.45, 4.20),
        ("2023-09-20", 3.45, 4.20),
        ("2023-08-21", 3.45, 4.20),
        ("2023-07-20", 3.55, 4.20),
        ("2023-06-20", 3.55, 4.20),
        ("2023-05-22", 3.65, 4.30),
        ("2023-04-20", 3.65, 4.30),
        ("2023-03-20", 3.65, 4.30),
        ("2023-02-20", 3.65, 4.30),
        ("2023-01-20", 3.65, 4.30),
        ("2022-12-20", 3.65, 4.30),
        ("2022-11-21", 3.65, 4.30),
        ("2022-10-20", 3.65, 4.30),
        ("2022-09-20", 3.65, 4.30),
        ("2022-08-22", 3.65, 4.30),
        ("2022-07-20", 3.70, 4.45),
        ("2022-06-20", 3.70, 4.45),
        ("2022-05-20", 3.70, 4.45),
        ("2022-04-20", 3.70, 4.60),
        ("2022-03-21", 3.70, 4.60),
        ("2022-02-21", 3.70, 4.60),
        ("2022-01-20", 3.70, 4.60),
        ("2021-12-20", 3.80, 4.65),
        ("2021-11-22", 3.85, 4.65),
        ("2021-10-20", 3.85, 4.65),
        ("2021-09-22", 3.85, 4.65),
        ("2021-08-20", 3.85, 4.65),
        ("2021-07-20", 3.85, 4.65),
        ("2021-06-21", 3.85, 4.65),
        ("2021-05-20", 3.85, 4.65),
        ("2021-04-20", 3.85, 4.65),
        ("2021-03-22", 3.85, 4.65),
        ("2021-02-22", 3.85, 4.65),
        ("2021-01-20", 3.85, 4.65),
        ("2020-12-21", 3.85, 4.65),
        ("2020-11-20", 3.85, 4.65),
        ("2020-10-20", 3.85, 4.65),
        ("2020-09-21", 3.85, 4.65),
        ("2020-08-20", 3.85, 4.65),
        ("2020-07-20", 3.85, 4.65),
        ("2020-06-22", 3.85, 4.65),
        ("2020-05-20", 3.85, 4.65),
        ("2020-04-20", 3.85, 4.65),
        ("2020-03-20", 4.05, 4.75),
        ("2020-02-20", 4.05, 4.75),
        ("2020-01-20", 4.15, 4.80),
        ("2019-12-20", 4.15, 4.80),
        ("2019-11-20", 4.15, 4.80),
        ("2019-10-21", 4.20, 4.85),
        ("2019-09-20", 4.20, 4.85),
        ("2019-08-20", 4.25, 4.85),
    ]


# LPR data - loaded from YAML or fallback to embedded
LPR_DATA = load_lpr_data_from_yaml()


class InterestCalculator:
    """
    Universal interest calculator for debt review.

    Supports multiple calculation methods required for bankruptcy debt analysis.
    """

    def __init__(self):
        """Initialize calculator with LPR data."""
        self.lpr_rates = {}
        for date_str, rate_1y, rate_5y in LPR_DATA:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            self.lpr_rates[date_obj] = {
                "1y": Decimal(str(rate_1y)),
                "5y": Decimal(str(rate_5y))
            }
        self.lpr_dates = sorted(self.lpr_rates.keys(), reverse=True)

    def _to_decimal(self, value: Union[float, int, str, Decimal]) -> Decimal:
        """Convert value to Decimal safely."""
        if isinstance(value, Decimal):
            return value
        return Decimal(str(value))

    def _parse_date(self, date_input: Union[datetime, str]) -> datetime:
        """Parse date from string or datetime."""
        if isinstance(date_input, datetime):
            return date_input
        try:
            return datetime.strptime(date_input, '%Y-%m-%d')
        except ValueError:
            return datetime.strptime(date_input, '%Y/%m/%d')

    def _get_lpr_rate(self, date: datetime, term: str = "1y") -> Decimal:
        """Get LPR rate for a given date."""
        for lpr_date in self.lpr_dates:
            if date >= lpr_date:
                return self.lpr_rates[lpr_date][term]
        # Return earliest rate if date is before all records
        return self.lpr_rates[self.lpr_dates[-1]][term]

    def calculate_simple_interest(
        self,
        principal: float,
        start_date: str,
        end_date: str,
        annual_rate: float,
        include_end_date: bool = True
    ) -> Dict[str, Any]:
        """
        Calculate simple interest.

        Formula: Interest = Principal × Rate × Days / 365

        Args:
            include_end_date: If True, count end_date as a full day (default: True)
                             This matches legal practice where "至X日止" includes X date.
        """
        principal_d = self._to_decimal(principal)
        rate_d = self._to_decimal(annual_rate) / 100
        start = self._parse_date(start_date)
        end = self._parse_date(end_date)

        # In legal practice, "自X日起至Y日止" includes both X and Y dates
        days = (end - start).days
        if include_end_date:
            days += 1

        if days <= 0:
            return {"error": "End date must be after start date"}

        interest = principal_d * rate_d * days / 365
        interest = interest.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        return {
            "calculation_type": "simple",
            "principal": float(principal_d),
            "annual_rate": annual_rate,
            "start_date": start_date,
            "end_date": end_date,
            "days": days,
            "interest": float(interest),
            "total": float(principal_d + interest),
            "formula": f"{principal} × {annual_rate}% × {days}/365 = {float(interest)}"
        }

    def calculate_lpr_interest(
        self,
        principal: float,
        start_date: str,
        end_date: str,
        multiplier: float = 1.0,
        lpr_term: str = "1y",
        include_end_date: bool = True
    ) -> Dict[str, Any]:
        """
        Calculate interest based on floating LPR rate.

        Rate changes with each LPR announcement.

        Args:
            principal: Principal amount
            start_date: Start date (YYYY-MM-DD)
            end_date: End date (YYYY-MM-DD)
            multiplier: LPR multiplier (e.g., 1.5 for 150% LPR)
            lpr_term: "1y" for 1-year LPR, "5y" for 5-year LPR
            include_end_date: If True, count end_date as a full day (default: True)
                             This matches legal practice where "至X日止" includes X date.
        """
        principal_d = self._to_decimal(principal)
        multiplier_d = self._to_decimal(multiplier)
        start = self._parse_date(start_date)
        end = self._parse_date(end_date)

        if end <= start:
            return {"error": "End date must be after start date"}

        # Adjust end date if we need to include it as a full day
        # In legal practice, "自X日起至Y日止" includes both X and Y dates
        if include_end_date:
            end = end + timedelta(days=1)

        # Get LPR dates in ascending order (from old to new) for proper iteration
        lpr_dates_asc = sorted(self.lpr_dates)

        # Calculate interest for each LPR period
        periods = []
        total_interest = Decimal('0')
        current_date = start

        while current_date < end:
            # Find the next LPR change date (in chronological order)
            next_lpr_date = None
            for lpr_date in lpr_dates_asc:
                if lpr_date > current_date and lpr_date < end:
                    next_lpr_date = lpr_date
                    break

            period_end = next_lpr_date if next_lpr_date else end
            days = (period_end - current_date).days

            if days > 0:
                base_rate = self._get_lpr_rate(current_date, lpr_term)
                effective_rate = base_rate * multiplier_d
                period_interest = principal_d * effective_rate / 100 * days / 365
                period_interest = period_interest.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

                periods.append({
                    "start": current_date.strftime('%Y-%m-%d'),
                    "end": period_end.strftime('%Y-%m-%d'),
                    "days": days,
                    "lpr_rate": float(base_rate),
                    "effective_rate": float(effective_rate),
                    "interest": float(period_interest)
                })

                total_interest += period_interest

            current_date = period_end

        total_days = (end - start).days

        return {
            "calculation_type": "lpr",
            "principal": float(principal_d),
            "multiplier": multiplier,
            "lpr_term": lpr_term,
            "start_date": start_date,
            "end_date": end_date,
            "total_days": total_days,
            "interest": float(total_interest),
            "total": float(principal_d + total_interest),
            "periods": periods
        }

    def calculate_delay_interest(
        self,
        principal: float,
        start_date: str,
        end_date: str,
        include_end_date: bool = True
    ) -> Dict[str, Any]:
        """
        Calculate delayed performance interest (迟延履行期间的加倍部分债务利息).

        根据《民事诉讼法》第260条及最高人民法院司法解释规定：
        迟延履行期间的加倍部分债务利息 = 本金 × 日利率万分之1.75 × 天数

        注意：这是固定日利率，不是LPR浮动利率！
        日利率万分之1.75 = 0.0175% = 0.000175

        Args:
            include_end_date: If True, count end_date as a full day (default: True)
                             This matches legal practice where "至X日止" includes X date.
        """
        principal_d = self._to_decimal(principal)
        start = self._parse_date(start_date)
        end = self._parse_date(end_date)

        # In legal practice, "自X日起至Y日止" includes both X and Y dates
        days = (end - start).days
        if include_end_date:
            days += 1

        if days <= 0:
            return {"error": "End date must be after start date"}

        # 固定日利率万分之1.75
        daily_rate = Decimal('0.000175')  # 0.0175% per day

        interest = principal_d * daily_rate * days
        interest = interest.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        return {
            "calculation_type": "delay",
            "principal": float(principal_d),
            "daily_rate": "万分之1.75 (0.0175%)",
            "daily_rate_decimal": float(daily_rate),
            "start_date": start_date,
            "end_date": end_date,
            "days": days,
            "interest": float(interest),
            "total": float(principal_d + interest),
            "formula": f"{principal} × 0.000175 × {days} = {float(interest)}",
            "legal_basis": "《民事诉讼法》第260条"
        }

    def calculate_penalty_interest(
        self,
        principal: float,
        start_date: str,
        end_date: str,
        annual_rate: float,
        penalty_cap: Optional[float] = None
    ) -> Dict[str, Any]:
        """
        Calculate penalty interest with optional cap.

        Penalty interest is capped at 24% annually per judicial interpretation.
        """
        # Apply 24% cap if rate exceeds
        capped_rate = min(annual_rate, 24.0)

        result = self.calculate_simple_interest(
            principal=principal,
            start_date=start_date,
            end_date=end_date,
            annual_rate=capped_rate
        )

        result["calculation_type"] = "penalty"
        result["original_rate"] = annual_rate
        result["capped_rate"] = capped_rate
        result["rate_was_capped"] = annual_rate > 24.0

        # Apply additional penalty cap if specified
        if penalty_cap and result["interest"] > penalty_cap:
            result["interest"] = penalty_cap
            result["total"] = float(self._to_decimal(principal) + self._to_decimal(penalty_cap))
            result["penalty_cap_applied"] = True

        return result

    def calculate_share_ratio(
        self,
        total_amount: float,
        share_ratio: float,
        description: str = ""
    ) -> Dict[str, Any]:
        """
        计算银团贷款份额比例金额。

        用于银团贷款案件中，按份额比例分配总金额。

        Args:
            total_amount: 银团总金额（本金或利息）
            share_ratio: 份额比例（如13.95表示13.95%）
            description: 描述信息（如"判决确认银团总利息"）

        Returns:
            包含份额计算结果的字典
        """
        total_d = self._to_decimal(total_amount)
        ratio_d = self._to_decimal(share_ratio) / 100  # 转换为小数

        share_amount = total_d * ratio_d
        share_amount = share_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        return {
            "calculation_type": "share_ratio",
            "total_amount": float(total_d),
            "share_ratio": share_ratio,
            "share_ratio_decimal": float(ratio_d),
            "share_amount": float(share_amount),
            "description": description,
            "formula": f"{total_amount:,.2f} × {share_ratio}% = {float(share_amount):,.2f}"
        }

    def calculate_confirmed_amount(
        self,
        confirmed_amount: float,
        source: str = "",
        description: str = ""
    ) -> Dict[str, Any]:
        """
        处理判决书确认的固定金额（直接使用，不计算）。

        用于判决书已明确确认的金额，如"截至2018年5月15日利息为247,674,737.97元"。

        Args:
            confirmed_amount: 判决确认的金额
            source: 来源（如"(2018)鄂72民初1270号民事判决书"）
            description: 描述信息（如"截至2018年5月15日的利息"）

        Returns:
            包含确认金额信息的字典
        """
        amount_d = self._to_decimal(confirmed_amount)
        amount_d = amount_d.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        return {
            "calculation_type": "confirmed",
            "confirmed_amount": float(amount_d),
            "source": source,
            "description": description,
            "note": "此金额由判决书直接确认，无需计算"
        }

    def apply_max_limit(
        self,
        calculated_total: float,
        max_limit: float,
        description: str = ""
    ) -> Dict[str, Any]:
        """
        应用最高额担保限额封顶。

        当计算总额超过最高额担保限额时，按限额封顶。

        Args:
            calculated_total: 计算得出的总额（本金+利息+费用）
            max_limit: 最高额担保限额
            description: 描述信息（如"最高额保证1.5亿元"）

        Returns:
            包含封顶检查结果的字典
        """
        total_d = self._to_decimal(calculated_total)
        limit_d = self._to_decimal(max_limit)

        exceeds_limit = total_d > limit_d
        final_amount = min(total_d, limit_d)
        final_amount = final_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        excess_amount = (total_d - limit_d) if exceeds_limit else Decimal('0')

        return {
            "calculation_type": "max_limit",
            "calculated_total": float(total_d),
            "max_limit": float(limit_d),
            "exceeds_limit": exceeds_limit,
            "final_amount": float(final_amount),
            "excess_amount": float(excess_amount) if exceeds_limit else 0,
            "description": description,
            "note": f"{'计算总额超过最高额限额，按限额封顶' if exceeds_limit else '计算总额未超过最高额限额'}",
            "formula": f"min({calculated_total:,.2f}, {max_limit:,.2f}) = {float(final_amount):,.2f}"
        }

    def calculate(
        self,
        calculation_type: str,
        principal: float = 0,
        start_date: str = "",
        end_date: str = "",
        rate: Optional[float] = None,
        multiplier: float = 1.0,
        lpr_term: str = "1y",
        **kwargs
    ) -> Dict[str, Any]:
        """
        Universal calculation entry point.

        Args:
            calculation_type: One of "simple", "lpr", "delay", "compound", "penalty",
                            "share_ratio", "confirmed", "max_limit"
            principal: Principal amount (for interest calculations)
            start_date: Start date (YYYY-MM-DD) (for interest calculations)
            end_date: End date (YYYY-MM-DD) (for interest calculations)
            rate: Annual rate for simple/penalty calculations
            multiplier: LPR multiplier (default 1.0)
            lpr_term: "1y" or "5y" for LPR calculations

        Additional kwargs for new calculation types:
            share_ratio: For SHARE_RATIO - total_amount, share_ratio, description
            confirmed: For CONFIRMED - confirmed_amount, source, description
            max_limit: For MAX_LIMIT - calculated_total, max_limit, description

        Returns:
            Calculation result dictionary
        """
        try:
            calc_type = CalculationType(calculation_type.lower())

            if calc_type == CalculationType.SIMPLE:
                if rate is None:
                    return {"error": "Rate is required for simple interest"}
                return self.calculate_simple_interest(principal, start_date, end_date, rate)

            elif calc_type == CalculationType.LPR:
                return self.calculate_lpr_interest(principal, start_date, end_date, multiplier, lpr_term)

            elif calc_type == CalculationType.DELAY:
                return self.calculate_delay_interest(principal, start_date, end_date)

            elif calc_type == CalculationType.PENALTY:
                if rate is None:
                    return {"error": "Rate is required for penalty interest"}
                return self.calculate_penalty_interest(
                    principal, start_date, end_date, rate,
                    penalty_cap=kwargs.get("penalty_cap")
                )

            elif calc_type == CalculationType.COMPOUND:
                # Compound interest calculation (simplified)
                if rate is None:
                    return {"error": "Rate is required for compound interest"}
                # For now, treat as simple interest
                # TODO: Implement proper compound interest
                return self.calculate_simple_interest(principal, start_date, end_date, rate)

            # ===== 新增计算类型 =====
            elif calc_type == CalculationType.SHARE_RATIO:
                # 份额比例计算（银团贷款）
                total_amount = kwargs.get("total_amount")
                share_ratio = kwargs.get("share_ratio")
                if total_amount is None or share_ratio is None:
                    return {"error": "total_amount and share_ratio are required for share_ratio calculation"}
                return self.calculate_share_ratio(
                    total_amount=total_amount,
                    share_ratio=share_ratio,
                    description=kwargs.get("description", "")
                )

            elif calc_type == CalculationType.CONFIRMED:
                # 判决确认金额（直接使用）
                confirmed_amount = kwargs.get("confirmed_amount")
                if confirmed_amount is None:
                    return {"error": "confirmed_amount is required for confirmed calculation"}
                return self.calculate_confirmed_amount(
                    confirmed_amount=confirmed_amount,
                    source=kwargs.get("source", ""),
                    description=kwargs.get("description", "")
                )

            elif calc_type == CalculationType.MAX_LIMIT:
                # 最高额限额封顶
                calculated_total = kwargs.get("calculated_total")
                max_limit = kwargs.get("max_limit")
                if calculated_total is None or max_limit is None:
                    return {"error": "calculated_total and max_limit are required for max_limit calculation"}
                return self.apply_max_limit(
                    calculated_total=calculated_total,
                    max_limit=max_limit,
                    description=kwargs.get("description", "")
                )

        except ValueError as e:
            return {"error": f"Invalid calculation type: {calculation_type}"}
        except Exception as e:
            logger.error(f"Calculation error: {e}")
            return {"error": str(e)}


# Singleton instance
_calculator = None


def get_calculator() -> InterestCalculator:
    """Get or create calculator instance."""
    global _calculator
    if _calculator is None:
        _calculator = InterestCalculator()
    return _calculator


def reload_lpr_data():
    """
    Reload LPR data from YAML file.

    Call this function when LPR rates are updated to refresh the calculator.
    """
    global LPR_DATA, _calculator
    LPR_DATA = load_lpr_data_from_yaml()
    _calculator = None  # Reset singleton to pick up new data
    logger.info("LPR data reloaded")


def calculate_interest(
    calculation_type: str,
    principal: float,
    start_date: str,
    end_date: str,
    **kwargs
) -> Dict[str, Any]:
    """
    Convenience function for calculating interest.

    This is the main entry point for the calculator tool.
    """
    calculator = get_calculator()
    return calculator.calculate(
        calculation_type=calculation_type,
        principal=principal,
        start_date=start_date,
        end_date=end_date,
        **kwargs
    )


class ExcelExporter:
    """
    Export calculation results to Excel with detailed process breakdown.

    输出格式匹配原方案，包含：
    - 汇总表：各项计算结果汇总
    - LPR利息明细：分段LPR利率计算过程
    - 迟延履行利息明细：固定日利率计算过程
    """

    def __init__(self):
        if not EXCEL_SUPPORT:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        # Define styles
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        self.title_font = Font(bold=True, size=14)
        self.money_format = '#,##0.00'
        self.rate_format = '0.00%'
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export_calculation_results(
        self,
        results: List[Dict[str, Any]],
        creditor_name: str,
        output_path: str,
        claimed_amounts: Optional[Dict[str, float]] = None
    ) -> str:
        """
        Export calculation results to Excel file.

        Args:
            results: List of calculation result dictionaries
            creditor_name: Creditor name for file naming
            output_path: Path to save the Excel file
            claimed_amounts: Optional dict of claimed amounts for comparison
                             {"利息损失": 342449.99, "迟延履行利息": 278500.92}

        Returns:
            Path to the created Excel file
        """
        wb = Workbook()

        # Create summary sheet
        self._create_summary_sheet(wb, results, creditor_name, claimed_amounts)

        # Create detail sheets for each calculation type
        lpr_results = [r for r in results if r.get('calculation_type') == 'lpr']
        delay_results = [r for r in results if r.get('calculation_type') == 'delay']
        simple_results = [r for r in results if r.get('calculation_type') == 'simple']

        if lpr_results:
            self._create_lpr_detail_sheet(wb, lpr_results)

        if delay_results:
            self._create_delay_detail_sheet(wb, delay_results)

        if simple_results:
            self._create_simple_detail_sheet(wb, simple_results)

        # Save file
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_file))

        logger.info(f"Excel calculation file saved to: {output_file}")
        return str(output_file)

    def _create_summary_sheet(
        self,
        wb: Workbook,
        results: List[Dict],
        creditor_name: str,
        claimed_amounts: Optional[Dict[str, float]] = None
    ):
        """Create summary sheet with all calculation results."""
        ws = wb.active
        ws.title = "计算汇总"

        # Title
        ws['A1'] = f"{creditor_name} - 利息计算汇总表"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')

        # Headers
        headers = ['序号', '计算类型', '本金(元)', '计算期间', '天数', '计算结果(元)', '申报金额(元)', '差异(元)', '确认金额(元)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        calc_type_names = {
            'lpr': 'LPR浮动利息',
            'delay': '迟延履行利息',
            'simple': '固定利率利息',
            'penalty': '罚息',
            'compound': '复利'
        }

        claimed_map = {
            'lpr': claimed_amounts.get('利息损失', None) if claimed_amounts else None,
            'delay': claimed_amounts.get('迟延履行利息', None) if claimed_amounts else None,
        }

        # Deduplicate results (keep unique combinations)
        seen = set()
        unique_results = []
        for r in results:
            key = (r.get('calculation_type'), r.get('principal'), r.get('start_date'), r.get('end_date'))
            if key not in seen:
                seen.add(key)
                unique_results.append(r)

        for idx, result in enumerate(unique_results, 1):
            calc_type = result.get('calculation_type', 'unknown')
            calc_name = calc_type_names.get(calc_type, calc_type)
            principal = result.get('principal', 0)
            start_date = result.get('start_date', '')
            end_date = result.get('end_date', '')
            days = result.get('days', result.get('total_days', 0))
            interest = result.get('interest', 0)

            claimed = claimed_map.get(calc_type)
            if claimed is not None:
                diff = interest - claimed
                confirmed = min(interest, claimed)  # 就低原则
            else:
                diff = None
                confirmed = interest

            row = idx + 3
            ws.cell(row=row, column=1, value=idx).border = self.thin_border
            ws.cell(row=row, column=2, value=calc_name).border = self.thin_border

            cell = ws.cell(row=row, column=3, value=principal)
            cell.number_format = self.money_format
            cell.border = self.thin_border

            ws.cell(row=row, column=4, value=f"{start_date} 至 {end_date}").border = self.thin_border
            ws.cell(row=row, column=5, value=days).border = self.thin_border

            cell = ws.cell(row=row, column=6, value=interest)
            cell.number_format = self.money_format
            cell.border = self.thin_border

            if claimed is not None:
                cell = ws.cell(row=row, column=7, value=claimed)
                cell.number_format = self.money_format
                cell.border = self.thin_border

                cell = ws.cell(row=row, column=8, value=diff)
                cell.number_format = self.money_format
                cell.border = self.thin_border
            else:
                ws.cell(row=row, column=7, value='-').border = self.thin_border
                ws.cell(row=row, column=8, value='-').border = self.thin_border

            cell = ws.cell(row=row, column=9, value=confirmed)
            cell.number_format = self.money_format
            cell.border = self.thin_border

        # Auto-adjust column widths
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions['D'].width = 28

    def _create_lpr_detail_sheet(self, wb: Workbook, lpr_results: List[Dict]):
        """Create LPR interest detail sheet with period breakdown."""
        ws = wb.create_sheet("LPR利息明细")

        # Title
        ws['A1'] = "LPR浮动利率利息计算明细"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:H1')

        current_row = 3

        for result in lpr_results:
            # Calculation parameters
            ws.cell(row=current_row, column=1, value="本金(元):")
            ws.cell(row=current_row, column=2, value=result.get('principal', 0)).number_format = self.money_format
            ws.cell(row=current_row, column=3, value="LPR倍数:")
            ws.cell(row=current_row, column=4, value=result.get('multiplier', 1))
            ws.cell(row=current_row, column=5, value="LPR期限:")
            ws.cell(row=current_row, column=6, value=result.get('lpr_term', '1y'))
            current_row += 1

            ws.cell(row=current_row, column=1, value="计算期间:")
            ws.cell(row=current_row, column=2, value=f"{result.get('start_date', '')} 至 {result.get('end_date', '')}")
            ws.cell(row=current_row, column=3, value="总天数:")
            ws.cell(row=current_row, column=4, value=result.get('total_days', 0))
            current_row += 2

            # Period detail headers
            headers = ['序号', '计息起日', '计息止日', '天数', 'LPR基准(%)', '执行利率(%)', '期间利息(元)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.thin_border
            current_row += 1

            # Period details
            periods = result.get('periods', [])
            for idx, period in enumerate(periods, 1):
                ws.cell(row=current_row, column=1, value=idx).border = self.thin_border
                ws.cell(row=current_row, column=2, value=period.get('start', '')).border = self.thin_border
                ws.cell(row=current_row, column=3, value=period.get('end', '')).border = self.thin_border
                ws.cell(row=current_row, column=4, value=period.get('days', 0)).border = self.thin_border

                cell = ws.cell(row=current_row, column=5, value=period.get('lpr_rate', 0))
                cell.border = self.thin_border

                cell = ws.cell(row=current_row, column=6, value=period.get('effective_rate', 0))
                cell.border = self.thin_border

                cell = ws.cell(row=current_row, column=7, value=period.get('interest', 0))
                cell.number_format = self.money_format
                cell.border = self.thin_border
                current_row += 1

            # Total
            ws.cell(row=current_row, column=6, value="合计:").font = self.header_font
            cell = ws.cell(row=current_row, column=7, value=result.get('interest', 0))
            cell.number_format = self.money_format
            cell.font = self.header_font
            current_row += 3

        # Auto-adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_delay_detail_sheet(self, wb: Workbook, delay_results: List[Dict]):
        """Create delay interest detail sheet."""
        ws = wb.create_sheet("迟延履行利息明细")

        # Title
        ws['A1'] = "迟延履行期间加倍部分债务利息计算明细"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')

        # Legal basis
        ws['A3'] = "法律依据: 《民事诉讼法》第260条"
        ws['A4'] = "计算公式: 本金 × 日利率万分之1.75 × 天数"
        ws['A5'] = "日利率: 0.0175% (万分之1.75)"

        current_row = 7

        # Headers
        headers = ['序号', '本金(元)', '计息起日', '计息止日', '天数', '利息(元)', '计算公式']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        current_row += 1

        # Data
        for idx, result in enumerate(delay_results, 1):
            ws.cell(row=current_row, column=1, value=idx).border = self.thin_border

            cell = ws.cell(row=current_row, column=2, value=result.get('principal', 0))
            cell.number_format = self.money_format
            cell.border = self.thin_border

            ws.cell(row=current_row, column=3, value=result.get('start_date', '')).border = self.thin_border
            ws.cell(row=current_row, column=4, value=result.get('end_date', '')).border = self.thin_border
            ws.cell(row=current_row, column=5, value=result.get('days', 0)).border = self.thin_border

            cell = ws.cell(row=current_row, column=6, value=result.get('interest', 0))
            cell.number_format = self.money_format
            cell.border = self.thin_border

            ws.cell(row=current_row, column=7, value=result.get('formula', '')).border = self.thin_border
            current_row += 1

        # Auto-adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions['G'].width = 35

    def _create_simple_detail_sheet(self, wb: Workbook, simple_results: List[Dict]):
        """Create simple interest detail sheet."""
        ws = wb.create_sheet("固定利率利息明细")

        # Title
        ws['A1'] = "固定利率利息计算明细"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:G1')

        current_row = 3

        # Headers
        headers = ['序号', '本金(元)', '年利率(%)', '计息起日', '计息止日', '天数', '利息(元)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        current_row += 1

        # Data
        for idx, result in enumerate(simple_results, 1):
            ws.cell(row=current_row, column=1, value=idx).border = self.thin_border

            cell = ws.cell(row=current_row, column=2, value=result.get('principal', 0))
            cell.number_format = self.money_format
            cell.border = self.thin_border

            ws.cell(row=current_row, column=3, value=result.get('annual_rate', 0)).border = self.thin_border
            ws.cell(row=current_row, column=4, value=result.get('start_date', '')).border = self.thin_border
            ws.cell(row=current_row, column=5, value=result.get('end_date', '')).border = self.thin_border
            ws.cell(row=current_row, column=6, value=result.get('days', 0)).border = self.thin_border

            cell = ws.cell(row=current_row, column=7, value=result.get('interest', 0))
            cell.number_format = self.money_format
            cell.border = self.thin_border
            current_row += 1

        # Auto-adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15


def export_to_excel(
    results: List[Dict[str, Any]],
    creditor_name: str,
    output_path: str,
    claimed_amounts: Optional[Dict[str, float]] = None
) -> str:
    """
    Export calculation results to Excel file.

    Convenience function for exporting calculation results.

    Args:
        results: List of calculation result dictionaries
        creditor_name: Creditor name
        output_path: Output file path
        claimed_amounts: Optional dict of claimed amounts for comparison

    Returns:
        Path to the created Excel file
    """
    if not EXCEL_SUPPORT:
        logger.warning("Excel export not available. Install openpyxl: pip install openpyxl")
        return ""

    exporter = ExcelExporter()
    return exporter.export_calculation_results(
        results=results,
        creditor_name=creditor_name,
        output_path=output_path,
        claimed_amounts=claimed_amounts
    )
