#!/usr/bin/env python3
"""
通用债权利息计算器 - CLI版本
支持单利、LPR浮动利率、迟延履行加倍利息、复利计算
单文件，无外部依赖，结构化输入输出
"""

import sys
import json
import argparse
import re
import csv
import os
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP, getcontext
from typing import Dict, List, Tuple, Optional, Union
from calendar import monthrange

# 设置Decimal全局精度配置 - 用于金融计算的高精度要求
getcontext().prec = 28  # 28位有效数字，足够处理大金额计算
getcontext().rounding = ROUND_HALF_UP  # 统一使用四舍五入规则
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# 内嵌LPR数据（从TSV文件提取的核心数据）
LPR_DATA = [
    ("2025-07-21", 3.00, 3.50),
    ("2025-06-20", 3.00, 3.50),
    ("2025-05-20", 3.00, 3.50),
    ("2025-04-21", 3.10, 3.60),
    ("2025-03-20", 3.10, 3.60),
    ("2025-02-20", 3.10, 3.60),
    ("2025-01-20", 3.10, 3.60),
    ("2024-12-20", 3.10, 3.60),
    ("2024-11-20", 3.10, 3.60),
    ("2024-10-21", 3.10, 3.60),
    ("2024-09-20", 3.35, 3.85),
    ("2024-08-20", 3.35, 3.85),
    ("2024-07-22", 3.35, 3.85),
    ("2024-06-20", 3.45, 3.95),
    ("2024-05-20", 3.45, 3.95),
    ("2024-04-22", 3.45, 3.95),
    ("2024-03-20", 3.45, 3.95),
    ("2024-02-20", 3.45, 3.95),
    ("2024-01-22", 3.45, 4.20),
    ("2023-12-20", 3.45, 4.20),
    ("2023-11-20", 3.45, 4.20),
    ("2023-10-20", 3.45, 4.20),
    ("2023-09-20", 3.45, 4.20),
    ("2023-08-21", 3.45, 4.20),
    ("2023-07-20", 3.55, 4.20),
    ("2023-06-20", 3.55, 4.20),
    ("2023-05-22", 3.65, 4.30),
    ("2023-04-20", 3.65, 4.30),
    ("2023-03-20", 3.65, 4.30),
    ("2023-02-20", 3.65, 4.30),
    ("2023-01-20", 3.65, 4.30),
    ("2022-12-20", 3.65, 4.30),
    ("2022-11-21", 3.65, 4.30),
    ("2022-10-20", 3.65, 4.30),
    ("2022-09-20", 3.65, 4.30),
    ("2022-08-22", 3.65, 4.30),
    ("2022-07-20", 3.70, 4.45),
    ("2022-06-20", 3.70, 4.45),
    ("2022-05-20", 3.70, 4.45),
    ("2022-04-20", 3.70, 4.60),
    ("2022-03-21", 3.70, 4.60),
    ("2022-02-21", 3.70, 4.60),
    ("2022-01-20", 3.70, 4.60),
    ("2021-12-20", 3.80, 4.65),
    ("2021-11-22", 3.85, 4.65),
    ("2021-10-20", 3.85, 4.65),
    ("2021-09-22", 3.85, 4.65),
    ("2021-08-20", 3.85, 4.65),
    ("2021-07-20", 3.85, 4.65),
    ("2021-06-21", 3.85, 4.65),
    ("2021-05-20", 3.85, 4.65),
    ("2021-04-20", 3.85, 4.65),
    ("2021-03-22", 3.85, 4.65),
    ("2021-02-22", 3.85, 4.65),
    ("2021-01-20", 3.85, 4.65),
    ("2020-12-21", 3.85, 4.65),
    ("2020-11-20", 3.85, 4.65),
    ("2020-10-20", 3.85, 4.65),
    ("2020-09-21", 3.85, 4.65),
    ("2020-08-20", 3.85, 4.65),
    ("2020-07-20", 3.85, 4.65),
    ("2020-06-22", 3.85, 4.65),
    ("2020-05-20", 3.85, 4.65),
    ("2020-04-20", 3.85, 4.65),
    ("2020-03-20", 4.05, 4.75),
    ("2020-02-20", 4.05, 4.75),
    ("2020-01-20", 4.15, 4.80),
    ("2019-12-20", 4.15, 4.80),
    ("2019-11-20", 4.15, 4.80),
    ("2019-10-21", 4.20, 4.85),
    ("2019-09-20", 4.20, 4.85),
    ("2019-08-20", 4.25, 4.85),
]


class UniversalDebtCalculatorCLI:
    """通用债权利息计算器CLI版本"""
    
    def __init__(self):
        """初始化计算器"""
        # 解析LPR数据到字典，按日期排序，使用Decimal存储利率确保精度
        self.lpr_rates = {}
        for date_str, rate_1y, rate_5y in LPR_DATA:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            self.lpr_rates[date_obj] = {
                "1年期": self.to_decimal(rate_1y),
                "5年期以上": self.to_decimal(rate_5y)
            }
        
        # 按日期排序（最新的在前）
        self.lpr_dates = sorted(self.lpr_rates.keys(), reverse=True)
    
    def parse_date(self, date_input: Union[datetime, str]) -> datetime:
        """解析日期输入"""
        if isinstance(date_input, datetime):
            return date_input
        elif isinstance(date_input, str):
            try:
                return datetime.strptime(date_input, '%Y-%m-%d')
            except ValueError:
                try:
                    return datetime.strptime(date_input, '%Y/%m/%d')
                except ValueError:
                    raise ValueError(f"无效的日期格式: {date_input}")
        else:
            raise ValueError(f"不支持的日期类型: {type(date_input)}")
    
    def to_decimal(self, value: Union[float, int, str, Decimal]) -> Decimal:
        """将输入值安全转换为Decimal，确保精度"""
        if isinstance(value, Decimal):
            return value
        elif isinstance(value, (int, float, str)):
            try:
                return Decimal(str(value))
            except (ValueError, TypeError):
                raise ValueError(f"无法转换为Decimal: {value}")
        else:
            raise ValueError(f"不支持的数值类型: {type(value)}")
    
    def validate_parameters(self, principal: Union[float, Decimal], start_date: datetime, end_date: datetime,
                          initial_accumulated_interest: Union[float, Decimal] = 0.0) -> Tuple[bool, str]:
        """验证计算参数"""
        if principal <= 0:
            return False, "本金必须大于0"
        
        if start_date > end_date:
            return False, "起息日不能晚于停息日"
        
        if end_date > datetime.now():
            return False, "停息日不能晚于当前日期"
        
        if start_date.year < 1900:
            return False, "起息日过于久远，请核实"
        
        if initial_accumulated_interest < 0:
            return False, "起息日已产生利息不能为负数"
        
        return True, "参数验证通过"
    
    def round_amount(self, amount: Union[float, Decimal]) -> Decimal:
        """金额四舍五入到分，返回Decimal确保精度"""
        if isinstance(amount, Decimal):
            return amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            return self.to_decimal(amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    def format_decimal_display(self, value: Union[float, Decimal], decimal_places: int = 2) -> str:
        """
        格式化Decimal值用于显示，去除不必要的尾随零
        
        Args:
            value: 要格式化的数值
            decimal_places: 保留的小数位数，默认2位（用于金额）
            
        Returns:
            格式化后的字符串
        """
        if isinstance(value, Decimal):
            decimal_val = value
        else:
            decimal_val = self.to_decimal(value)
        
        # 根据数值类型决定格式化方式
        if decimal_places == 2:
            # 金额格式：保留2位小数
            formatted = f"{decimal_val:.2f}".rstrip('0').rstrip('.')
        elif decimal_places == 6:
            # 利率格式：最多保留6位小数，去除尾随零
            formatted = f"{decimal_val:.6f}".rstrip('0').rstrip('.')
        else:
            # 通用格式：保留指定小数位，去除尾随零
            format_str = f"{{:.{decimal_places}f}}"
            formatted = format_str.format(decimal_val).rstrip('0').rstrip('.')
        
        # 如果结果为空（全是零），返回"0"
        if not formatted or formatted == '':
            formatted = '0'
            
        return formatted
    
    def error_result(self, error_msg: str) -> Dict:
        """生成错误结果"""
        return {
            'success': False,
            'error': True,
            'error_message': error_msg,
            'total_interest': 0.0,
            'total_compound_interest': 0.0,
            'status_code': 1
        }
    
    def success_result(self, data: Dict) -> Dict:
        """生成成功结果"""
        data.update({
            'success': True,
            'error': False,
            'status_code': 0
        })
        return data
    
    # ==================== 单利计算相关方法 ====================
    
    def calculate_simple_interest(self, principal: Union[float, Decimal], start_date: Union[datetime, str], 
                                end_date: Union[datetime, str], annual_rate: Union[float, Decimal] = None, 
                                daily_rate: Union[float, Decimal] = None, base_days: int = 360) -> Dict:
        """
        计算普通单利利息
        
        Args:
            principal: 本金
            start_date: 起息日
            end_date: 停息日
            annual_rate: 年利率（百分比），与daily_rate二选一
            daily_rate: 日利率（如万分之三则输入0.03），与annual_rate二选一
            base_days: 基准天数（360或365）
        
        Returns:
            计算结果字典
        """
        try:
            # 验证利率参数（必须有且只有一个）
            if annual_rate is None and daily_rate is None:
                return self.error_result("必须提供年利率或日利率之一")
            if annual_rate is not None and daily_rate is not None:
                return self.error_result("年利率和日利率只能提供其中之一")
            
            # 转换为Decimal确保精度
            principal_decimal = self.to_decimal(principal)
            base_days_decimal = self.to_decimal(base_days)
            
            # 日期格式转换
            start_date = self.parse_date(start_date)
            end_date = self.parse_date(end_date)
            
            # 参数验证
            is_valid, error_msg = self.validate_parameters(principal_decimal, start_date, end_date)
            if not is_valid:
                return self.error_result(error_msg)
            
            # 计算天数
            days = (end_date - start_date).days + 1
            days_decimal = self.to_decimal(days)
            
            # 根据提供的利率类型计算利息
            if daily_rate is not None:
                # 使用日利率计算：利息 = 本金 × 日利率 × 天数
                daily_rate_decimal = self.to_decimal(daily_rate) / Decimal('100')  # 转换为小数形式
                interest_decimal = principal_decimal * daily_rate_decimal * days_decimal
                rate_info = f"日利率{daily_rate}%（即万分之{daily_rate*100:.1f}）"
                # 删除年化利率转换，避免365天和360天基数不一致的误差
            else:
                # 使用年利率计算：利息 = 本金 × (天数/基准天数) × (年利率/100)
                annual_rate_decimal = self.to_decimal(annual_rate)
                rate_fraction = annual_rate_decimal / Decimal('100')
                days_fraction = days_decimal / base_days_decimal
                interest_decimal = principal_decimal * days_fraction * rate_fraction
                rate_info = f"年利率{annual_rate}%"
            
            interest = self.round_amount(interest_decimal)
            
            # 构建结果
            if daily_rate is not None:
                # 日利率计算的结果
                daily_rate_display = self.to_decimal(daily_rate)
                result = {
                    'calculation_type': 'simple_interest',
                    'calculation_type_name': '普通单利（日利率）',
                    'principal': float(principal_decimal),
                    'start_date': start_date.strftime('%Y-%m-%d'),
                    'end_date': end_date.strftime('%Y-%m-%d'),
                    'days': days,
                    'daily_rate': float(daily_rate_display),
                    'rate_info': rate_info,
                    'total_interest': float(interest),
                    'formula': f"利息 = {self.format_decimal_display(principal_decimal)} × {days} × {self.format_decimal_display(daily_rate_display, 6)}% = {self.format_decimal_display(interest)}",
                    'segments': [{
                        'period': f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}",
                        'days': days,
                        'rate': float(daily_rate_display),
                        'rate_desc': rate_info,
                        'interest': float(interest),
                        'calculation': f"{self.format_decimal_display(principal_decimal)} × {days} × {self.format_decimal_display(daily_rate_display, 6)}% = {self.format_decimal_display(interest)}"
                    }]
                }
            else:
                # 年利率计算的结果
                result = {
                    'calculation_type': 'simple_interest',
                    'calculation_type_name': '普通单利',
                    'principal': float(principal_decimal),
                    'start_date': start_date.strftime('%Y-%m-%d'),
                    'end_date': end_date.strftime('%Y-%m-%d'),
                    'days': days,
                    'annual_rate': float(annual_rate_decimal),
                    'base_days': base_days,
                    'total_interest': float(interest),
                    'formula': f"利息 = {self.format_decimal_display(principal_decimal)} × ({days}/{base_days}) × {self.format_decimal_display(annual_rate_decimal, 6)}% = {self.format_decimal_display(interest)}",
                    'segments': [{
                        'period': f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}",
                        'days': days,
                        'rate': float(annual_rate_decimal),
                        'rate_desc': f"{self.format_decimal_display(annual_rate_decimal, 6)}%",
                        'interest': float(interest),
                        'calculation': f"{self.format_decimal_display(principal_decimal)} × {days} ÷ {base_days} × {self.format_decimal_display(annual_rate_decimal, 6)}% = {self.format_decimal_display(interest)}"
                    }]
                }
            
            return self.success_result(result)
            
        except Exception as e:
            return self.error_result(f"计算错误: {str(e)}")
    
    # ==================== LPR计算相关方法 ====================
    
    def get_lpr_rate_for_date(self, target_date: datetime, lpr_term: str) -> Optional[Decimal]:
        """获取指定日期的LPR利率"""
        # 找到不晚于目标日期的最新LPR利率
        for lpr_date in self.lpr_dates:
            if lpr_date <= target_date:
                return self.lpr_rates[lpr_date].get(lpr_term)
        
        # 如果没有找到，使用最早的利率
        if self.lpr_dates:
            earliest_date = self.lpr_dates[-1]
            return self.lpr_rates[earliest_date].get(lpr_term)
        
        return None
    
    def get_lpr_segments(self, start_date: datetime, end_date: datetime, lpr_term: str) -> List[Dict]:
        """获取LPR分段数据"""
        segments = []
        current_date = start_date
        
        while current_date <= end_date:
            # 找到当前日期适用的LPR利率
            current_rate = self.get_lpr_rate_for_date(current_date, lpr_term)
            if current_rate is None:
                break
            
            # 找到下一个利率真正变动的日期
            next_change_date = end_date
            for lpr_date in sorted(self.lpr_dates):  # 从早到晚排序
                if lpr_date > current_date and lpr_date <= end_date:
                    # 检查这个日期的利率是否真的与当前利率不同
                    next_rate = self.get_lpr_rate_for_date(lpr_date, lpr_term)
                    if next_rate != current_rate:
                        next_change_date = lpr_date - timedelta(days=1)
                        break
            
            # 计算分段天数
            segment_end = min(next_change_date, end_date)
            segment_days = (segment_end - current_date).days + 1
            
            segments.append({
                'start_date': current_date,
                'end_date': segment_end,
                'days': segment_days,
                'lpr_rate': current_rate,
                'period_desc': f"{current_date.strftime('%Y-%m-%d')} 至 {segment_end.strftime('%Y-%m-%d')}"
            })
            
            # 移动到下一段
            current_date = segment_end + timedelta(days=1)
        
        return segments
    
    def calculate_lpr_floating_interest(self, principal: Union[float, Decimal], start_date: Union[datetime, str],
                                      end_date: Union[datetime, str], lpr_multiplier: Union[float, Decimal] = 1.0,
                                      lpr_term: str = '1年期', base_days: int = 360) -> Dict:
        """
        计算LPR浮动利率利息
        
        Args:
            principal: 本金
            start_date: 起息日
            end_date: 停息日
            lpr_multiplier: LPR倍数
            lpr_term: LPR期限类型
            base_days: 基准天数
        
        Returns:
            计算结果字典
        """
        try:
            # 转换为Decimal确保精度
            principal_decimal = self.to_decimal(principal)
            lpr_multiplier_decimal = self.to_decimal(lpr_multiplier)
            base_days_decimal = self.to_decimal(base_days)
            
            # 日期格式转换
            start_date = self.parse_date(start_date)
            end_date = self.parse_date(end_date)
            
            # 参数验证
            is_valid, error_msg = self.validate_parameters(principal_decimal, start_date, end_date)
            if not is_valid:
                return self.error_result(error_msg)
            
            # 获取LPR分段数据
            segments_data = self.get_lpr_segments(start_date, end_date, lpr_term)
            
            if not segments_data:
                return self.error_result("无法获取LPR利率数据")
            
            # 计算各分段利息
            segments = []
            total_interest_decimal = Decimal('0')
            total_days = 0
            
            for segment in segments_data:
                seg_start = segment['start_date']
                seg_end = segment['end_date']
                seg_days = segment['days']
                lpr_rate_decimal = segment['lpr_rate']  # 已经是Decimal类型
                actual_rate_decimal = lpr_rate_decimal * lpr_multiplier_decimal
                
                # 计算分段利息（使用Decimal确保精度）
                seg_days_decimal = self.to_decimal(seg_days)
                rate_fraction = actual_rate_decimal / Decimal('100')
                days_fraction = seg_days_decimal / base_days_decimal
                seg_interest_decimal = principal_decimal * days_fraction * rate_fraction
                seg_interest = self.round_amount(seg_interest_decimal)
                
                segments.append({
                    'period': segment['period_desc'],
                    'days': seg_days,
                    'lpr_rate': float(lpr_rate_decimal),
                    'multiplier': float(lpr_multiplier_decimal),
                    'actual_rate': float(actual_rate_decimal),
                    'rate_desc': f"LPR {self.format_decimal_display(lpr_rate_decimal, 6)}% × {self.format_decimal_display(lpr_multiplier_decimal, 6)} = {self.format_decimal_display(actual_rate_decimal, 6)}%",
                    'interest': float(seg_interest),
                    'calculation': f"{self.format_decimal_display(principal_decimal)} × {seg_days} ÷ {base_days} × {self.format_decimal_display(actual_rate_decimal, 6)}% = {self.format_decimal_display(seg_interest)}"
                })
                
                total_interest_decimal += seg_interest
                total_days += seg_days
            
            total_interest = self.round_amount(total_interest_decimal)
            
            # 构建结果
            result = {
                'calculation_type': 'lpr_floating_interest',
                'calculation_type_name': 'LPR浮动利率',
                'principal': float(principal_decimal),
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'total_days': total_days,
                'lpr_multiplier': float(lpr_multiplier_decimal),
                'lpr_term': lpr_term,
                'base_days': base_days,
                'total_interest': float(total_interest),
                'segments_count': len(segments),
                'segments': segments
            }
            
            return self.success_result(result)
            
        except Exception as e:
            return self.error_result(f"计算错误: {str(e)}")
    
    # ==================== 迟延履行加倍利息计算 ====================
    
    def calculate_delay_performance_interest(self, principal: Union[float, Decimal], start_date: Union[datetime, str],
                                           end_date: Union[datetime, str]) -> Dict:
        """
        计算迟延履行期间加倍利息
        
        Args:
            principal: 本金
            start_date: 起息日
            end_date: 停息日
        
        Returns:
            计算结果字典
        """
        try:
            # 转换为Decimal确保精度
            principal_decimal = self.to_decimal(principal)
            
            # 日期格式转换
            start_date = self.parse_date(start_date)
            end_date = self.parse_date(end_date)
            
            # 参数验证
            is_valid, error_msg = self.validate_parameters(principal_decimal, start_date, end_date)
            if not is_valid:
                return self.error_result(error_msg)
            
            # 计算天数
            days = (end_date - start_date).days + 1
            days_decimal = self.to_decimal(days)
            
            # 计算利息 - 固定日利率万分之1.75（使用Decimal确保精度）
            daily_rate_decimal = Decimal('0.000175')  # 万分之1.75
            interest_decimal = principal_decimal * days_decimal * daily_rate_decimal
            interest = self.round_amount(interest_decimal)
            
            # 构建结果
            result = {
                'calculation_type': 'delay_performance_interest',
                'calculation_type_name': '迟延履行加倍利息',
                'principal': float(principal_decimal),
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'days': days,
                'daily_rate': float(daily_rate_decimal),
                'daily_rate_desc': '万分之1.75',
                'total_interest': float(interest),
                'formula': f"利息 = {self.format_decimal_display(principal_decimal)} × {days} × 0.0175% = {self.format_decimal_display(interest)}",
                'segments': [{
                    'period': f"{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}",
                    'days': days,
                    'rate': float(daily_rate_decimal * Decimal('100')),
                    'rate_desc': '万分之1.75',
                    'interest': float(interest),
                    'calculation': f"{self.format_decimal_display(principal_decimal)} × {days} × 0.0175% = {self.format_decimal_display(interest)}"
                }]
            }
            
            return self.success_result(result)
            
        except Exception as e:
            return self.error_result(f"计算错误: {str(e)}")
    
    # ==================== 复利计算相关方法 ====================
    
    def parse_settlement_cycle(self, cycle_str: str) -> Dict:
        """
        解析结息周期字符串
        
        支持的格式：
        - "每月20日" -> 每月的20号
        - "每月末" -> 每月的最后一天
        - "每季末" -> 每季度末（3月31日、6月30日、9月30日、12月31日）
        - "每半年末" -> 每半年末（6月30日、12月31日）
        - "每年末" -> 每年末（12月31日）
        - "每30天" -> 每30天固定周期
        """
        cycle_str = cycle_str.strip()
        
        # 每月XX日
        month_day_pattern = r'每月(\d{1,2})日'
        match = re.match(month_day_pattern, cycle_str)
        if match:
            day = int(match.group(1))
            if 1 <= day <= 31:
                return {'type': 'monthly_day', 'day': day}
            else:
                raise ValueError(f"无效的日期: {day}日")
        
        # 每月末
        if cycle_str == '每月末':
            return {'type': 'monthly_end'}
        
        # 每季末
        if cycle_str == '每季末':
            return {'type': 'quarterly_end'}
        
        # 每半年末
        if cycle_str == '每半年末':
            return {'type': 'semiannual_end'}
        
        # 每年末
        if cycle_str == '每年末':
            return {'type': 'annual_end'}
        
        # 每XX天
        days_pattern = r'每(\d+)天'
        match = re.match(days_pattern, cycle_str)
        if match:
            days = int(match.group(1))
            if days > 0:
                return {'type': 'fixed_days', 'days': days}
            else:
                raise ValueError(f"无效的天数: {days}")
        
        raise ValueError(f"不支持的结息周期格式: {cycle_str}")
    
    def generate_settlement_dates(self, start_date: datetime, end_date: datetime, cycle_config: Dict) -> List[datetime]:
        """根据结息周期配置生成结息日期列表"""
        dates = []
        cycle_type = cycle_config['type']
        
        if cycle_type == 'monthly_day':
            day = cycle_config['day']
            current_date = start_date
            
            while current_date <= end_date:
                # 找到当前月份的结息日
                year = current_date.year
                month = current_date.month
                
                # 处理特殊情况：如果指定日期超过当月天数，使用当月最后一天
                max_day = monthrange(year, month)[1]
                actual_day = min(day, max_day)
                
                settlement_date = datetime(year, month, actual_day)
                
                # 如果结息日在起息日之后且在停息日之前或等于
                if settlement_date > start_date and settlement_date <= end_date:
                    dates.append(settlement_date)
                
                # 移动到下一个月
                if month == 12:
                    current_date = datetime(year + 1, 1, 1)
                else:
                    current_date = datetime(year, month + 1, 1)
        
        elif cycle_type == 'monthly_end':
            current_date = start_date
            
            while current_date <= end_date:
                year = current_date.year
                month = current_date.month
                
                # 当月最后一天
                max_day = monthrange(year, month)[1]
                settlement_date = datetime(year, month, max_day)
                
                if settlement_date > start_date and settlement_date <= end_date:
                    dates.append(settlement_date)
                
                # 移动到下一个月
                if month == 12:
                    current_date = datetime(year + 1, 1, 1)
                else:
                    current_date = datetime(year, month + 1, 1)
        
        elif cycle_type == 'quarterly_end':
            # 季末：3月31日、6月30日、9月30日、12月31日
            quarter_months = [3, 6, 9, 12]
            quarter_days = [31, 30, 30, 31]
            
            current_year = start_date.year
            end_year = end_date.year
            
            while current_year <= end_year:
                for i, month in enumerate(quarter_months):
                    settlement_date = datetime(current_year, month, quarter_days[i])
                    if settlement_date > start_date and settlement_date <= end_date:
                        dates.append(settlement_date)
                current_year += 1
        
        elif cycle_type == 'semiannual_end':
            # 半年末：6月30日、12月31日
            current_year = start_date.year
            end_year = end_date.year
            
            while current_year <= end_year:
                for month, day in [(6, 30), (12, 31)]:
                    settlement_date = datetime(current_year, month, day)
                    if settlement_date > start_date and settlement_date <= end_date:
                        dates.append(settlement_date)
                current_year += 1
        
        elif cycle_type == 'annual_end':
            # 年末：12月31日
            current_year = start_date.year
            end_year = end_date.year
            
            while current_year <= end_year:
                settlement_date = datetime(current_year, 12, 31)
                if settlement_date > start_date and settlement_date <= end_date:
                    dates.append(settlement_date)
                current_year += 1
        
        elif cycle_type == 'fixed_days':
            days = cycle_config['days']
            current_date = start_date + timedelta(days=days)
            
            while current_date <= end_date:
                dates.append(current_date)
                current_date += timedelta(days=days)
        
        return sorted(dates)
    
    def split_periods(self, start_date: datetime, end_date: datetime, settlement_cycle: str) -> List[Dict]:
        """根据结息周期分割计息期间"""
        try:
            cycle_config = self.parse_settlement_cycle(settlement_cycle)
        except ValueError as e:
            raise ValueError(f"结息周期解析错误: {e}")
        
        # 生成结息日期
        settlement_dates = self.generate_settlement_dates(start_date, end_date, cycle_config)
        
        # 构建期间列表
        periods = []
        current_start = start_date
        
        for settlement_date in settlement_dates:
            period_end = settlement_date
            days = (period_end - current_start).days + 1
            
            periods.append({
                'start_date': current_start,
                'end_date': period_end,
                'days': days,
                'period_desc': f"{current_start.strftime('%Y-%m-%d')} 至 {period_end.strftime('%Y-%m-%d')}"
            })
            
            current_start = period_end + timedelta(days=1)
        
        # 处理最后一个期间（如果有剩余）
        if current_start <= end_date:
            days = (end_date - current_start).days + 1
            periods.append({
                'start_date': current_start,
                'end_date': end_date,
                'days': days,
                'period_desc': f"{current_start.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}"
            })
        
        return periods
    
    def calculate_compound_interest(self, principal: Union[float, Decimal], start_date: Union[datetime, str],
                                  end_date: Union[datetime, str], annual_rate: Union[float, Decimal],
                                  settlement_cycle: str, base_days: int = 360,
                                  initial_accumulated_interest: Union[float, Decimal] = 0.0) -> Dict:
        """
        计算复利
        
        Args:
            principal: 本金
            start_date: 起息日
            end_date: 停息日
            annual_rate: 年利率（百分比）
            settlement_cycle: 结息周期
            base_days: 基准天数（360或365）
            initial_accumulated_interest: 起息日已产生的利息（用于复利计算）
        
        Returns:
            计算结果字典
        """
        try:
            # 转换为Decimal确保精度
            principal_decimal = self.to_decimal(principal)
            annual_rate_decimal = self.to_decimal(annual_rate)
            base_days_decimal = self.to_decimal(base_days)
            initial_accumulated_interest_decimal = self.to_decimal(initial_accumulated_interest)
            
            # 日期格式转换
            start_date = self.parse_date(start_date)
            end_date = self.parse_date(end_date)
            
            # 参数验证
            is_valid, error_msg = self.validate_parameters(principal_decimal, start_date, end_date, initial_accumulated_interest_decimal)
            if not is_valid:
                return self.error_result(error_msg)
            
            # 分割计息期间
            periods = self.split_periods(start_date, end_date, settlement_cycle)
            
            if not periods:
                return self.error_result("无法生成有效的计息期间")
            
            # 计算各期间利息和复利（全程使用Decimal确保精度）
            segments = []
            total_interest_decimal = Decimal('0')  # 总利息
            total_compound_interest_decimal = Decimal('0')  # 总复利
            accumulated_interest_decimal = initial_accumulated_interest_decimal  # 累积利息（用于计算复利，不包含复利）
            total_days = 0
            rate_fraction = annual_rate_decimal / Decimal('100')
            
            for i, period in enumerate(periods):
                period_start = period['start_date']
                period_end = period['end_date']
                period_days = period['days']
                period_days_decimal = self.to_decimal(period_days)
                
                # 计算本期利息：利息 = 本金 × 天数/基准天数 × 利率
                days_fraction = period_days_decimal / base_days_decimal
                period_interest_decimal = principal_decimal * days_fraction * rate_fraction
                period_interest = self.round_amount(period_interest_decimal)
                
                # 计算本期复利：复利 = 累积利息 × 天数/基准天数 × 利率
                period_compound_interest_decimal = accumulated_interest_decimal * days_fraction * rate_fraction
                period_compound_interest = self.round_amount(period_compound_interest_decimal)
                
                # 构建分段信息
                segment = {
                    'period_number': i + 1,
                    'period': period['period_desc'],
                    'start_date': period_start.strftime('%Y-%m-%d'),
                    'end_date': period_end.strftime('%Y-%m-%d'),
                    'days': period_days,
                    'rate': float(annual_rate_decimal),
                    'rate_desc': f"{self.format_decimal_display(annual_rate_decimal, 6)}%",
                    'interest': float(period_interest),
                    'compound_interest': float(period_compound_interest),
                    'accumulated_interest_before': float(accumulated_interest_decimal),
                    'total_period_amount': float(period_interest + period_compound_interest),
                    'interest_calculation': f"{self.format_decimal_display(principal_decimal)} × {period_days} ÷ {base_days} × {self.format_decimal_display(annual_rate_decimal, 6)}% = {self.format_decimal_display(period_interest)}",
                    'compound_calculation': f"{self.format_decimal_display(accumulated_interest_decimal)} × {period_days} ÷ {base_days} × {self.format_decimal_display(annual_rate_decimal, 6)}% = {self.format_decimal_display(period_compound_interest)}"
                }
                
                segments.append(segment)
                
                # 累加到总数（使用Decimal运算）
                total_interest_decimal += period_interest
                total_compound_interest_decimal += period_compound_interest
                total_days += period_days
                
                # 更新累积利息（只累加利息，不累加复利）
                accumulated_interest_decimal += period_interest
            
            total_interest = self.round_amount(total_interest_decimal)
            total_compound_interest = self.round_amount(total_compound_interest_decimal)
            total_amount = self.round_amount(total_interest + total_compound_interest)
            
            # 计算实际年化收益率
            effective_rate = Decimal('0')
            if total_days > 0 and principal_decimal > 0:
                total_days_decimal = self.to_decimal(total_days)
                rate_ratio = total_amount / principal_decimal
                annualized_ratio = rate_ratio * (base_days_decimal / total_days_decimal) * Decimal('100')
                effective_rate = self.round_amount(annualized_ratio)
            
            # 构建结果
            result = {
                'calculation_type': 'compound_interest',
                'calculation_type_name': '复利计算',
                'principal': float(principal_decimal),
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'total_days': total_days,
                'annual_rate': float(annual_rate_decimal),
                'settlement_cycle': settlement_cycle,
                'base_days': base_days,
                'initial_accumulated_interest': float(initial_accumulated_interest_decimal),
                'total_interest': float(total_interest),
                'total_compound_interest': float(total_compound_interest),
                'total_amount': float(total_amount),
                'periods_count': len(segments),
                'segments': segments,
                'summary': {
                    'principal': float(principal_decimal),
                    'initial_accumulated_interest': float(initial_accumulated_interest_decimal),
                    'total_interest': float(total_interest),
                    'total_compound_interest': float(total_compound_interest),
                    'total_amount': float(total_amount),
                    'effective_rate': float(effective_rate)
                }
            }
            
            return self.success_result(result)
            
        except Exception as e:
            return self.error_result(f"计算错误: {str(e)}")


def print_result_table(result: Dict):
    """以表格形式打印计算结果"""
    if result.get('error'):
        print(f"❌ 计算错误: {result.get('error_message', '未知错误')}")
        return
    
    calc_type = result.get('calculation_type')
    
    if calc_type == 'compound_interest':
        # 复利计算的特殊输出
        print_compound_result_table(result)
    else:
        # 单利、LPR、迟延履行的统一输出
        print_simple_result_table(result)


def print_simple_result_table(result: Dict):
    """以表格形式打印单利类计算结果"""
    print("\n" + "="*70)
    print("📊 债权利息计算结果")
    print("="*70)
    
    # 基本信息
    print(f"计算类型: {result.get('calculation_type_name', '未知类型')}")
    print(f"本金: {result.get('principal', 0):,.2f} 元")
    print(f"起息日: {result.get('start_date', '')}")
    print(f"停息日: {result.get('end_date', '')}")
    print(f"计息天数: {result.get('days', result.get('total_days', 0))} 天")
    
    # 特定信息
    calc_type = result.get('calculation_type')
    if calc_type == 'simple_interest':
        print(f"年利率: {result.get('annual_rate', 0)}%")
        print(f"基准天数: {result.get('base_days', 360)} 天/年")
    elif calc_type == 'lpr_floating_interest':
        print(f"LPR倍数: {result.get('lpr_multiplier', 1.0)}")
        print(f"LPR期限: {result.get('lpr_term', '1年期')}")
        print(f"基准天数: {result.get('base_days', 360)} 天/年")
        print(f"分段数量: {result.get('segments_count', 0)} 段")
    elif calc_type == 'delay_performance_interest':
        print(f"日利率: {result.get('daily_rate_desc', '万分之1.75')}")
    
    print("-" * 70)
    print(f"💰 利息总额: {result.get('total_interest', 0):,.2f} 元")
    print("-" * 70)
    
    # 分段详情（如果有多段）
    segments = result.get('segments', [])
    if len(segments) > 1:
        print("\n📈 分段计算详情:")
        for i, segment in enumerate(segments, 1):
            print(f"第{i}段: {segment['period']}")
            print(f"  天数: {segment['days']} 天")
            print(f"  利率: {segment['rate_desc']}")
            print(f"  利息: {segment['interest']:,.2f} 元")
    
    # 计算公式
    if 'formula' in result:
        print(f"\n📋 计算公式: {result['formula']}")
    
    print()


def print_compound_result_table(result: Dict):
    """以表格形式打印复利计算结果"""
    print("\n" + "="*80)
    print("💰 复利计算结果")
    print("="*80)
    
    # 基本信息
    print(f"计算类型: {result.get('calculation_type_name', '复利计算')}")
    print(f"本金: {result.get('principal', 0):,.2f} 元")
    print(f"起息日: {result.get('start_date', '')}")
    print(f"停息日: {result.get('end_date', '')}")
    print(f"计息天数: {result.get('total_days', 0)} 天")
    print(f"年利率: {result.get('annual_rate', 0)}%")
    print(f"结息周期: {result.get('settlement_cycle', '')}")
    print(f"基准天数: {result.get('base_days', 360)} 天/年")
    print(f"分期数量: {result.get('periods_count', 0)} 期")
    
    # 显示起息日已产生利息
    initial_interest = result.get('initial_accumulated_interest', 0)
    if initial_interest > 0:
        print(f"起息日已产生利息: {initial_interest:,.2f} 元")
    
    print("-" * 80)
    
    # 汇总信息
    summary = result.get('summary', {})
    print(f"💵 利息总额: {result.get('total_interest', 0):,.2f} 元")
    print(f"🔄 复利总额: {result.get('total_compound_interest', 0):,.2f} 元")
    print(f"📊 合计金额: {result.get('total_amount', 0):,.2f} 元")
    if 'effective_rate' in summary:
        print(f"📈 实际年化收益率: {summary['effective_rate']:.4f}%")
    
    print("-" * 80)
    
    # 分期详情
    segments = result.get('segments', [])
    if segments:
        print("\n📈 分期计算详情:")
        print(f"{'期数':<4} {'期间':<24} {'天数':<6} {'本期利息':<12} {'本期复利':<12} {'期间合计':<12}")
        print("-" * 80)
        
        for segment in segments:
            period_num = segment['period_number']
            period_desc = segment['period'][:22] + ".." if len(segment['period']) > 24 else segment['period']
            days = segment['days']
            interest = segment['interest']
            compound = segment['compound_interest']
            total_period = segment['total_period_amount']
            
            print(f"{period_num:<4} {period_desc:<24} {days:<6} {interest:>10,.2f} {compound:>10,.2f} {total_period:>10,.2f}")
        
        print("-" * 80)
        
        # 显示前几期的详细计算过程
        print("\n🔍 前3期详细计算过程:")
        for i, segment in enumerate(segments[:3]):
            print(f"\n第{segment['period_number']}期 ({segment['period']}):")
            print(f"  累积利息(期初): {segment['accumulated_interest_before']:,.2f} 元")
            print(f"  本期利息计算: {segment['interest_calculation']}")
            print(f"  本期复利计算: {segment['compound_calculation']}")
            print(f"  本期合计: {segment['interest']:,.2f} + {segment['compound_interest']:,.2f} = {segment['total_period_amount']:,.2f} 元")
    
    print()


def load_json_input(json_file: str) -> Dict:
    """从JSON文件加载输入参数"""
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"❌ 错误: JSON输入文件不存在: {json_file}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"❌ 错误: JSON文件格式错误: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ 错误: 无法读取JSON文件: {e}")
        sys.exit(1)


def save_json_output(result: Dict, json_file: str = None):
    """保存结果到JSON文件"""
    json_str = json.dumps(result, ensure_ascii=False, indent=2)
    
    if json_file:
        try:
            with open(json_file, 'w', encoding='utf-8') as f:
                f.write(json_str)
            print(f"✅ 结果已保存到JSON文件: {json_file}")
        except Exception as e:
            print(f"❌ 错误: 无法保存JSON文件: {e}")
            sys.exit(1)
    else:
        print(json_str)


def save_excel_output(result: Dict, excel_file: str, sheet_name: str = None, debtor_name: str = None, append_mode: bool = False):
    """
    保存结果到Excel文件（支持多sheet）
    
    Args:
        result: 计算结果字典
        excel_file: Excel文件路径
        sheet_name: Sheet名称
        debtor_name: 债权人名称
        append_mode: 是否追加模式（添加新sheet）
    """
    if not EXCEL_SUPPORT:
        print("❌ 错误: 未安装openpyxl库，无法生成Excel文件")
        print("请运行: pip install openpyxl")
        return
    
    if result.get('error'):
        print(f"❌ 无法保存Excel: {result.get('error_message', '计算错误')}")
        return
    
    try:
        # 判断是否需要加载现有文件
        if append_mode and os.path.exists(excel_file):
            wb = load_workbook(excel_file)
        else:
            wb = Workbook()
            # 删除默认的sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # 确定sheet名称
        calc_type_name = result.get('calculation_type_name', '计算结果')
        if not sheet_name:
            # 自动生成sheet名称
            existing_sheets = wb.sheetnames
            sheet_num = 1
            for s in existing_sheets:
                if calc_type_name in s:
                    sheet_num += 1
            sheet_name = f"{calc_type_name}_{sheet_num}" if sheet_num > 1 else calc_type_name
        
        # 创建新sheet
        if sheet_name in wb.sheetnames:
            # 如果sheet已存在，添加序号
            i = 1
            original_name = sheet_name
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name}_{i}"
                i += 1
        
        ws = wb.create_sheet(title=sheet_name)
        
        # 定义样式
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        row = 1
        
        # 写入标题
        ws.cell(row=row, column=1, value=f"===== 计算报告 =====").font = title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        
        ws.cell(row=row, column=1, value="生成时间")
        ws.cell(row=row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        row += 1
        
        if debtor_name:
            ws.cell(row=row, column=1, value="债权人")
            ws.cell(row=row, column=2, value=debtor_name)
            row += 1
        
        ws.cell(row=row, column=1, value="计算类型")
        ws.cell(row=row, column=2, value=calc_type_name)
        row += 2
        
        # 基本参数部分
        ws.cell(row=row, column=1, value="[基本参数]").font = header_font
        row += 1
        
        calc_type = result.get('calculation_type')
        
        # 基本参数表头
        ws.cell(row=row, column=1, value="参数名").font = header_font
        ws.cell(row=row, column=2, value="参数值").font = header_font
        for col in range(1, 3):
            ws.cell(row=row, column=col).fill = header_fill
        row += 1
        
        # 基本参数内容
        params = [
            ("本金", f"{result.get('principal', 0):,.2f}"),
            ("起息日", result.get('start_date', '')),
            ("停息日", result.get('end_date', ''))
        ]
        
        # 根据不同计算类型添加特定参数
        if calc_type == 'simple_interest':
            if result.get('daily_rate') is not None:
                # 日利率模式
                daily_rate_val = result.get('daily_rate', 0)
                params.extend([
                    ("日利率", f"{daily_rate_val}%（即万分之{daily_rate_val*100:.1f}）"),
                    ("计息天数", result.get('days', 0))
                ])
            else:
                # 年利率模式
                base_days_value = result.get('base_days', 360)
                base_days_source = result.get('_base_days_source', '')
                if base_days_source:
                    base_days_display = f"{base_days_value}（{base_days_source}）"
                else:
                    base_days_display = str(base_days_value)

                params.extend([
                    ("年利率", f"{result.get('annual_rate', 0)}%"),
                    ("基准天数", base_days_display),
                    ("计息天数", result.get('days', 0))
                ])
        elif calc_type == 'lpr_floating_interest':
            base_days_value = result.get('base_days', 360)
            base_days_source = result.get('_base_days_source', '')
            if base_days_source:
                base_days_display = f"{base_days_value}（{base_days_source}）"
            else:
                base_days_display = str(base_days_value)

            params.extend([
                ("LPR期限", result.get('lpr_term', '1年期')),
                ("LPR倍数", result.get('lpr_multiplier', 1.0)),
                ("基准天数", base_days_display),
                ("总天数", result.get('total_days', 0)),
                ("分段数量", result.get('segments_count', 0))
            ])
        elif calc_type == 'delay_performance_interest':
            params.extend([
                ("日利率", result.get('daily_rate_desc', '万分之1.75')),
                ("计息天数", result.get('days', 0))
            ])
        elif calc_type == 'compound_interest':
            base_days_value = result.get('base_days', 360)
            base_days_source = result.get('_base_days_source', '')
            if base_days_source:
                base_days_display = f"{base_days_value}（{base_days_source}）"
            else:
                base_days_display = str(base_days_value)

            params.extend([
                ("年利率", f"{result.get('annual_rate', 0)}%"),
                ("结息周期", result.get('settlement_cycle', '')),
                ("基准天数", base_days_display),
                ("总天数", result.get('total_days', 0)),
                ("分期数量", result.get('periods_count', 0))
            ])
            initial_interest = result.get('initial_accumulated_interest', 0)
            if initial_interest > 0:
                params.append(("起息日已产生利息", f"{initial_interest:,.2f}"))
        
        for param_name, param_value in params:
            ws.cell(row=row, column=1, value=param_name)
            ws.cell(row=row, column=2, value=str(param_value))
            row += 1
        
        row += 1
        
        # 计算明细部分
        segments = result.get('segments', [])
        if segments:
            ws.cell(row=row, column=1, value="[计算明细]").font = header_font
            row += 1
            
            # 根据计算类型设置不同的表头
            if calc_type == 'compound_interest':
                headers = ["期数", "期间", "天数", "利率", "本期利息", "本期复利", "累积利息(期初)", "计算公式(利息)", "计算公式(复利)"]
            elif calc_type == 'lpr_floating_interest':
                headers = ["期间", "天数", "LPR利率", "倍数", "实际利率", "利息", "计算公式"]
            else:
                headers = ["期间", "天数", "利率", "利息", "计算公式"]
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            row += 1
            
            # 写入明细数据
            for seg in segments:
                if calc_type == 'compound_interest':
                    values = [
                        seg.get('period_number', ''),
                        seg.get('period', ''),
                        seg.get('days', ''),
                        seg.get('rate_desc', ''),
                        f"{seg.get('interest', 0):,.2f}",
                        f"{seg.get('compound_interest', 0):,.2f}",
                        f"{seg.get('accumulated_interest_before', 0):,.2f}",
                        seg.get('interest_calculation', ''),
                        seg.get('compound_calculation', '')
                    ]
                elif calc_type == 'lpr_floating_interest':
                    values = [
                        seg.get('period', ''),
                        seg.get('days', ''),
                        f"{seg.get('lpr_rate', 0)}%",
                        seg.get('multiplier', ''),
                        seg.get('rate_desc', ''),
                        f"{seg.get('interest', 0):,.2f}",
                        seg.get('calculation', '')
                    ]
                else:
                    values = [
                        seg.get('period', ''),
                        seg.get('days', ''),
                        seg.get('rate_desc', ''),
                        f"{seg.get('interest', 0):,.2f}",
                        seg.get('calculation', '')
                    ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=str(value))
                    cell.border = border
                row += 1
        
        row += 1
        
        # 计算结果汇总
        ws.cell(row=row, column=1, value="[计算结果]").font = header_font
        row += 1
        
        ws.cell(row=row, column=1, value="项目").font = header_font
        ws.cell(row=row, column=2, value="金额(元)").font = header_font
        for col in range(1, 3):
            ws.cell(row=row, column=col).fill = header_fill
        row += 1
        
        if calc_type == 'compound_interest':
            results = [
                ("利息总额", f"{result.get('total_interest', 0):,.2f}"),
                ("复利总额", f"{result.get('total_compound_interest', 0):,.2f}"),
                ("合计金额", f"{result.get('total_amount', 0):,.2f}")
            ]
            summary = result.get('summary', {})
            if 'effective_rate' in summary:
                results.append(("实际年化收益率", f"{summary['effective_rate']:.4f}%"))
        else:
            results = [("利息总额", f"{result.get('total_interest', 0):,.2f}")]
        
        for item_name, item_value in results:
            ws.cell(row=row, column=1, value=item_name)
            ws.cell(row=row, column=2, value=item_value)
            row += 1
        
        # 添加计算公式说明（如果存在）
        if 'formula' in result:
            row += 1
            ws.cell(row=row, column=1, value="[计算公式]").font = header_font
            row += 1
            ws.cell(row=row, column=1, value=result['formula'])
            ws.merge_cells(f'A{row}:H{row}')
        
        # 调整列宽
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 保存文件
        wb.save(excel_file)
        print(f"✅ 结果已保存到Excel文件: {excel_file} (Sheet: {sheet_name})")
        
    except Exception as e:
        print(f"❌ 错误: 无法保存Excel文件: {e}")
        import traceback
        traceback.print_exc()


def save_csv_output(result: Dict, csv_file: str, debtor_name: str = None, append_mode: bool = False):
    """
    保存结果到CSV文件
    
    Args:
        result: 计算结果字典
        csv_file: CSV文件路径
        debtor_name: 债权人名称
        append_mode: 是否追加模式（用于同一债权人多笔计算）
    """
    if result.get('error'):
        print(f"❌ 无法保存CSV: {result.get('error_message', '计算错误')}")
        return
    
    try:
        # 判断是否需要追加模式
        file_exists = os.path.exists(csv_file)
        mode = 'a' if (append_mode and file_exists) else 'w'
        
        with open(csv_file, mode, encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            
            # 如果是追加模式，先添加分隔符
            if mode == 'a':
                writer.writerow([])
                writer.writerow([])
            
            # 写入计算编号和时间戳
            calc_number = 1
            if mode == 'a':
                # 简单计数，可以通过读取文件来获取准确的编号
                calc_number = 2  # 简化处理
            
            writer.writerow([f"===== 计算报告 #{calc_number} ====="])
            writer.writerow(["生成时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            if debtor_name:
                writer.writerow(["债权人", debtor_name])
            
            # 计算类型
            calc_type = result.get('calculation_type')
            calc_type_name = result.get('calculation_type_name', '未知类型')
            writer.writerow(["计算类型", calc_type_name])
            writer.writerow([])
            
            # 基本参数部分
            writer.writerow(["[基本参数]"])
            writer.writerow(["参数名", "参数值"])
            writer.writerow(["本金", f"{result.get('principal', 0):,.2f}"])
            writer.writerow(["起息日", result.get('start_date', '')])
            writer.writerow(["停息日", result.get('end_date', '')])
            
            # 根据不同计算类型输出特定参数
            if calc_type == 'simple_interest':
                writer.writerow(["年利率", f"{result.get('annual_rate', 0)}%"])
                writer.writerow(["基准天数", result.get('base_days', 360)])
                writer.writerow(["计息天数", result.get('days', 0)])
            elif calc_type == 'lpr_floating_interest':
                writer.writerow(["LPR期限", result.get('lpr_term', '1年期')])
                writer.writerow(["LPR倍数", result.get('lpr_multiplier', 1.0)])
                writer.writerow(["基准天数", result.get('base_days', 360)])
                writer.writerow(["总天数", result.get('total_days', 0)])
                writer.writerow(["分段数量", result.get('segments_count', 0)])
            elif calc_type == 'delay_performance_interest':
                writer.writerow(["日利率", result.get('daily_rate_desc', '万分之1.75')])
                writer.writerow(["计息天数", result.get('days', 0)])
            elif calc_type == 'compound_interest':
                writer.writerow(["年利率", f"{result.get('annual_rate', 0)}%"])
                writer.writerow(["结息周期", result.get('settlement_cycle', '')])
                writer.writerow(["基准天数", result.get('base_days', 360)])
                writer.writerow(["总天数", result.get('total_days', 0)])
                writer.writerow(["分期数量", result.get('periods_count', 0)])
                initial_interest = result.get('initial_accumulated_interest', 0)
                if initial_interest > 0:
                    writer.writerow(["起息日已产生利息", f"{initial_interest:,.2f}"])
            
            writer.writerow([])
            
            # 计算明细部分
            segments = result.get('segments', [])
            if segments:
                writer.writerow(["[计算明细]"])
                
                if calc_type == 'compound_interest':
                    # 复利计算的详细表格
                    writer.writerow(["期数", "期间", "天数", "利率", "本期利息", "本期复利", "累积利息(期初)", "计算公式(利息)", "计算公式(复利)"])
                    for seg in segments:
                        writer.writerow([
                            seg.get('period_number', ''),
                            seg.get('period', ''),
                            seg.get('days', ''),
                            seg.get('rate_desc', ''),
                            f"{seg.get('interest', 0):,.2f}",
                            f"{seg.get('compound_interest', 0):,.2f}",
                            f"{seg.get('accumulated_interest_before', 0):,.2f}",
                            seg.get('interest_calculation', ''),
                            seg.get('compound_calculation', '')
                        ])
                elif calc_type == 'lpr_floating_interest':
                    # LPR浮动利率的详细表格
                    writer.writerow(["期间", "天数", "LPR利率", "倍数", "实际利率", "利息", "计算公式"])
                    for seg in segments:
                        writer.writerow([
                            seg.get('period', ''),
                            seg.get('days', ''),
                            f"{seg.get('lpr_rate', 0)}%",
                            seg.get('multiplier', ''),
                            seg.get('rate_desc', ''),
                            f"{seg.get('interest', 0):,.2f}",
                            seg.get('calculation', '')
                        ])
                else:
                    # 单利和迟延履行的表格
                    writer.writerow(["期间", "天数", "利率", "利息", "计算公式"])
                    for seg in segments:
                        writer.writerow([
                            seg.get('period', ''),
                            seg.get('days', ''),
                            seg.get('rate_desc', ''),
                            f"{seg.get('interest', 0):,.2f}",
                            seg.get('calculation', '')
                        ])
            
            writer.writerow([])
            
            # 计算结果汇总
            writer.writerow(["[计算结果]"])
            writer.writerow(["项目", "金额(元)"])
            
            if calc_type == 'compound_interest':
                writer.writerow(["利息总额", f"{result.get('total_interest', 0):,.2f}"])
                writer.writerow(["复利总额", f"{result.get('total_compound_interest', 0):,.2f}"])
                writer.writerow(["合计金额", f"{result.get('total_amount', 0):,.2f}"])
                summary = result.get('summary', {})
                if 'effective_rate' in summary:
                    writer.writerow(["实际年化收益率", f"{summary['effective_rate']:.4f}%"])
            else:
                writer.writerow(["利息总额", f"{result.get('total_interest', 0):,.2f}"])
            
            # 添加计算公式说明（如果存在）
            if 'formula' in result:
                writer.writerow([])
                writer.writerow(["[计算公式]"])
                writer.writerow([result['formula']])
            
            print(f"✅ 结果已保存到CSV文件: {csv_file}")
            
    except Exception as e:
        print(f"❌ 错误: 无法保存CSV文件: {e}")
        sys.exit(1)


def apply_scenario_to_base_days(args, base_days_explicitly_set=False):
    """
    根据场景类型自动设置基准天数，并在未指定场景时发出警告

    优先级：explicit --base-days > --scenario > default 360

    Args:
        args: 命令行参数对象
        base_days_explicitly_set: 是否显式设置了base_days参数
    """
    # 场景到基准天数的映射
    SCENARIO_BASE_DAYS_MAP = {
        'judicial': 365,     # 司法裁判场景：法院判决、调解、仲裁 → 民法规定365天
        'financial': 360,    # 金融合同场景：银行贷款、金融机构 → 金融惯例360天
        'commercial': 365    # 商事合同场景：企业间合同 → 民法规定365天
    }

    # 场景名称映射
    scenario_names = {
        'judicial': '司法裁判',
        'financial': '金融合同',
        'commercial': '商事合同'
    }

    # 情况1：未指定scenario参数
    if not hasattr(args, 'scenario') or not args.scenario:
        # 如果也没有显式指定base_days，发出警告
        if not base_days_explicitly_set:
            print(f"⚠️  警告：未指定 --scenario 参数，将使用默认基准天数 360 天（金融惯例）")
            print(f"⚠️  建议：")
            print(f"   • 司法裁判场景（判决/调解/仲裁）：使用 --scenario judicial 或 --base-days 365")
            print(f"   • 金融合同场景（银行贷款）：使用 --scenario financial 或 --base-days 360")
            print(f"   • 商事合同场景（企业间合同）：使用 --scenario commercial 或 --base-days 365")
            print()
            # 记录参数来源（用于Excel输出）
            args._base_days_source = "默认值（金融惯例）"
            args._scenario_used = None
        else:
            # 用户显式指定了base_days，尊重用户选择
            args._base_days_source = "用户指定"
            args._scenario_used = None
        return

    # 情况2：指定了scenario参数
    scenario_base_days = SCENARIO_BASE_DAYS_MAP.get(args.scenario)

    if scenario_base_days is None:
        return  # 未知场景，不处理

    scenario_name = scenario_names.get(args.scenario, args.scenario)

    # 如果用户显式指定了base_days参数（通过命令行--base-days）
    if base_days_explicitly_set:
        # 检查是否与场景设置一致
        if args.base_days != scenario_base_days:
            print(f"⚠️  警告：--scenario={args.scenario} 建议使用{scenario_base_days}天，但您指定了 --base-days={args.base_days}")
            print(f"⚠️  将使用您显式指定的基准天数 {args.base_days} 天")
            # 记录：用户指定优先，但与场景冲突
            args._base_days_source = f"用户指定（与场景{scenario_name}冲突）"
            args._scenario_used = args.scenario
        else:
            # 场景和用户指定一致
            args._base_days_source = f"用户指定（与场景{scenario_name}一致）"
            args._scenario_used = args.scenario
        return

    # 用户只指定了scenario，没有显式指定base_days
    # 使用scenario设置base_days
    args.base_days = scenario_base_days

    # 记录参数来源
    args._base_days_source = f"场景：{scenario_name}"
    args._scenario_used = args.scenario

    if scenario_base_days == 365:
        print(f"ℹ️  根据场景'{scenario_name}'自动设置基准天数为 {scenario_base_days} 天（民法规定）")
    else:
        print(f"ℹ️  根据场景'{scenario_name}'自动设置基准天数为 {scenario_base_days} 天（金融惯例）")


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='通用债权利息计算器 - CLI版本',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 普通单利计算（使用场景参数自动设置基准天数）
  python universal_debt_calculator_cli.py simple --principal 100000 --start-date 2024-01-01 --end-date 2024-12-31 --rate 4.35 --scenario judicial

  # 普通单利计算（显式指定基准天数）
  python universal_debt_calculator_cli.py simple --principal 100000 --start-date 2024-01-01 --end-date 2024-12-31 --rate 4.35 --base-days 365

  # LPR浮动利率计算（金融合同场景）
  python universal_debt_calculator_cli.py lpr --principal 100000 --start-date 2024-01-01 --end-date 2024-12-31 --multiplier 1.5 --scenario financial

  # 迟延履行加倍利息计算
  python universal_debt_calculator_cli.py delay --principal 100000 --start-date 2024-01-01 --end-date 2024-12-31

  # 复利计算（商事合同场景）
  python universal_debt_calculator_cli.py compound --principal 100000 --start-date 2024-01-01 --end-date 2024-12-31 --rate 4.35 --cycle "每月末" --scenario commercial
  
  # 带起息日已产生利息的复利计算
  python universal_debt_calculator_cli.py compound --principal 100000 --start-date 2024-01-01 --end-date 2024-12-31 --rate 4.35 --cycle "每月末" --initial-interest 5000
  
  # JSON输入输出
  python universal_debt_calculator_cli.py simple --json-input input.json --json-output result.json
  
  # CSV输出（单笔计算）
  python universal_debt_calculator_cli.py simple --principal 100000 --start-date 2024-01-01 --end-date 2024-12-31 --rate 4.35 --csv-output result.csv --debtor "张三"
  
  # CSV输出（同一债权人多笔计算，追加模式）
  python universal_debt_calculator_cli.py simple --principal 50000 --start-date 2024-06-01 --end-date 2024-12-31 --rate 3.85 --csv-output result.csv --debtor "张三" --append
  
  # Excel输出（支持多sheet）
  python universal_debt_calculator_cli.py simple --principal 100000 --start-date 2024-01-01 --end-date 2024-12-31 --rate 4.35 --excel-output result.xlsx --sheet-name "违约金" --debtor "张三"
  python universal_debt_calculator_cli.py delay --principal 100000 --start-date 2024-01-01 --end-date 2024-12-31 --excel-output result.xlsx --sheet-name "迟延履行" --debtor "张三" --append

支持的计算类型:
  simple   : 普通单利计算
  lpr      : LPR浮动利率计算  
  delay    : 迟延履行加倍利息计算
  compound : 复利计算

支持的结息周期格式（复利计算）:
  - "每月20日" : 每月的20号结息
  - "每月末"   : 每月最后一天结息
  - "每季末"   : 每季度末结息(3/6/9/12月末)
  - "每半年末" : 每半年末结息(6/12月末)
  - "每年末"   : 每年末结息(12月末)
  - "每30天"   : 每30天固定周期结息
        """
    )
    
    # 子命令
    subparsers = parser.add_subparsers(dest='command', help='计算类型')
    
    # 普通单利计算
    simple_parser = subparsers.add_parser('simple', help='普通单利计算')
    simple_parser.add_argument('--principal', type=float, help='本金，单位：元')
    simple_parser.add_argument('--start-date', help='起息日，格式：YYYY-MM-DD')
    simple_parser.add_argument('--end-date', help='停息日，格式：YYYY-MM-DD')
    # 创建互斥组，年利率和日利率只能选择其一
    rate_group = simple_parser.add_mutually_exclusive_group()
    rate_group.add_argument('--rate', type=float, help='年利率，单位：百分比')
    rate_group.add_argument('--daily-rate', type=float, help='日利率，单位：百分比（如万分之三输入0.03）')
    simple_parser.add_argument('--base-days', type=int, default=360, help='基准天数，默认360')
    simple_parser.add_argument('--scenario', choices=['judicial', 'financial', 'commercial'],
                               help='场景类型（自动设置基准天数）：judicial(司法裁判,365天) | financial(金融合同,360天) | commercial(商事合同,365天)')
    simple_parser.add_argument('--json-input', help='JSON输入文件')
    simple_parser.add_argument('--json-output', nargs='?', const=True, help='JSON输出，可选指定文件名')
    simple_parser.add_argument('--csv-output', help='CSV输出文件')
    simple_parser.add_argument('--excel-output', help='Excel输出文件')
    simple_parser.add_argument('--sheet-name', help='Excel Sheet名称')
    simple_parser.add_argument('--debtor', help='债权人名称')
    simple_parser.add_argument('--append', action='store_true', help='追加模式（CSV追加内容，Excel添加新sheet）')
    
    # LPR浮动利率计算
    lpr_parser = subparsers.add_parser('lpr', help='LPR浮动利率计算')
    lpr_parser.add_argument('--principal', type=float, help='本金，单位：元')
    lpr_parser.add_argument('--start-date', help='起息日，格式：YYYY-MM-DD')
    lpr_parser.add_argument('--end-date', help='停息日，格式：YYYY-MM-DD')
    lpr_parser.add_argument('--multiplier', type=float, default=1.0, help='LPR倍数，默认1.0')
    lpr_parser.add_argument('--term', choices=['1年期', '5年期以上'], default='1年期', help='LPR期限')
    lpr_parser.add_argument('--base-days', type=int, default=360, help='基准天数，默认360')
    lpr_parser.add_argument('--scenario', choices=['judicial', 'financial', 'commercial'],
                            help='场景类型（自动设置基准天数）：judicial(司法裁判,365天) | financial(金融合同,360天) | commercial(商事合同,365天)')
    lpr_parser.add_argument('--json-input', help='JSON输入文件')
    lpr_parser.add_argument('--json-output', nargs='?', const=True, help='JSON输出，可选指定文件名')
    lpr_parser.add_argument('--csv-output', help='CSV输出文件')
    lpr_parser.add_argument('--excel-output', help='Excel输出文件')
    lpr_parser.add_argument('--sheet-name', help='Excel Sheet名称')
    lpr_parser.add_argument('--debtor', help='债权人名称')
    lpr_parser.add_argument('--append', action='store_true', help='追加模式（CSV追加内容，Excel添加新sheet）')
    
    # 迟延履行加倍利息计算
    delay_parser = subparsers.add_parser('delay', help='迟延履行加倍利息计算')
    delay_parser.add_argument('--principal', type=float, help='本金，单位：元')
    delay_parser.add_argument('--start-date', help='起息日，格式：YYYY-MM-DD')
    delay_parser.add_argument('--end-date', help='停息日，格式：YYYY-MM-DD')
    delay_parser.add_argument('--json-input', help='JSON输入文件')
    delay_parser.add_argument('--json-output', nargs='?', const=True, help='JSON输出，可选指定文件名')
    delay_parser.add_argument('--csv-output', help='CSV输出文件')
    delay_parser.add_argument('--excel-output', help='Excel输出文件')
    delay_parser.add_argument('--sheet-name', help='Excel Sheet名称')
    delay_parser.add_argument('--debtor', help='债权人名称')
    delay_parser.add_argument('--append', action='store_true', help='追加模式（CSV追加内容，Excel添加新sheet）')
    
    # 复利计算
    compound_parser = subparsers.add_parser('compound', help='复利计算')
    compound_parser.add_argument('--principal', type=float, help='本金，单位：元')
    compound_parser.add_argument('--start-date', help='起息日，格式：YYYY-MM-DD')
    compound_parser.add_argument('--end-date', help='停息日，格式：YYYY-MM-DD')
    compound_parser.add_argument('--rate', type=float, help='年利率，单位：百分比')
    compound_parser.add_argument('--cycle', help='结息周期，如"每月20日"、"每季末"等')
    compound_parser.add_argument('--base-days', type=int, default=360, help='基准天数，默认360')
    compound_parser.add_argument('--scenario', choices=['judicial', 'financial', 'commercial'],
                                 help='场景类型（自动设置基准天数）：judicial(司法裁判,365天) | financial(金融合同,360天) | commercial(商事合同,365天)')
    compound_parser.add_argument('--initial-interest', type=float, default=0.0, help='起息日已产生的利息，默认0')
    compound_parser.add_argument('--json-input', help='JSON输入文件')
    compound_parser.add_argument('--json-output', nargs='?', const=True, help='JSON输出，可选指定文件名')
    compound_parser.add_argument('--csv-output', help='CSV输出文件')
    compound_parser.add_argument('--excel-output', help='Excel输出文件')
    compound_parser.add_argument('--sheet-name', help='Excel Sheet名称')
    compound_parser.add_argument('--debtor', help='债权人名称')
    compound_parser.add_argument('--append', action='store_true', help='追加模式（CSV追加内容，Excel添加新sheet）')
    
    # 全局选项
    parser.add_argument('--version', action='version', version='通用债权利息计算器 CLI v1.0')
    
    # 如果没有参数，显示帮助信息
    if len(sys.argv) == 1:
        parser.print_help()
        return
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return
    
    # 创建计算器实例
    calculator = UniversalDebtCalculatorCLI()
    
    try:
        # 处理JSON输入
        if hasattr(args, 'json_input') and args.json_input:
            input_data = load_json_input(args.json_input)
            # 将JSON参数合并到args中
            for key, value in input_data.items():
                if not hasattr(args, key.replace('-', '_')) or getattr(args, key.replace('-', '_')) is None:
                    setattr(args, key.replace('-', '_'), value)
        
        # 执行相应的计算
        result = None
        
        if args.command == 'simple':
            # 检查必需参数和利率参数（二选一）
            if not all([args.principal, args.start_date, args.end_date]):
                print("❌ 错误: 普通单利计算需要提供 --principal, --start-date, --end-date 参数")
                sys.exit(1)
            if not (args.rate or args.daily_rate):
                print("❌ 错误: 普通单利计算需要提供 --rate（年利率）或 --daily-rate（日利率）参数之一")
                sys.exit(1)

            # 根据场景类型自动设置基准天数
            # 检测用户是否显式指定了--base-days参数
            base_days_explicitly_set = '--base-days' in sys.argv
            apply_scenario_to_base_days(args, base_days_explicitly_set)

            # 调用计算函数，传入相应的利率参数
            result = calculator.calculate_simple_interest(
                principal=args.principal,
                start_date=args.start_date,
                end_date=args.end_date,
                annual_rate=args.rate,
                daily_rate=args.daily_rate,
                base_days=args.base_days
            )
            
        elif args.command == 'lpr':
            if not all([args.principal, args.start_date, args.end_date]):
                print("❌ 错误: LPR浮动利率计算需要提供 --principal, --start-date, --end-date 参数")
                sys.exit(1)

            # 根据场景类型自动设置基准天数
            base_days_explicitly_set = '--base-days' in sys.argv
            apply_scenario_to_base_days(args, base_days_explicitly_set)

            result = calculator.calculate_lpr_floating_interest(
                args.principal, args.start_date, args.end_date, args.multiplier, args.term, args.base_days
            )
            
        elif args.command == 'delay':
            if not all([args.principal, args.start_date, args.end_date]):
                print("❌ 错误: 迟延履行加倍利息计算需要提供 --principal, --start-date, --end-date 参数")
                sys.exit(1)
            
            result = calculator.calculate_delay_performance_interest(
                args.principal, args.start_date, args.end_date
            )
            
        elif args.command == 'compound':
            if not all([args.principal, args.start_date, args.end_date, args.rate, args.cycle]):
                print("❌ 错误: 复利计算需要提供 --principal, --start-date, --end-date, --rate, --cycle 参数")
                sys.exit(1)

            # 根据场景类型自动设置基准天数
            base_days_explicitly_set = '--base-days' in sys.argv
            apply_scenario_to_base_days(args, base_days_explicitly_set)

            result = calculator.calculate_compound_interest(
                args.principal, args.start_date, args.end_date, args.rate, args.cycle, 
                args.base_days, args.initial_interest
            )
        
        # 输出结果
        if result:
            # 添加参数来源元数据到结果中（用于Excel输出）
            if hasattr(args, '_base_days_source'):
                result['_base_days_source'] = args._base_days_source
            if hasattr(args, '_scenario_used'):
                result['_scenario_used'] = args._scenario_used

            # Excel输出
            if hasattr(args, 'excel_output') and args.excel_output:
                debtor_name = getattr(args, 'debtor', None)
                sheet_name = getattr(args, 'sheet_name', None)
                append_mode = getattr(args, 'append', False)
                save_excel_output(result, args.excel_output, sheet_name, debtor_name, append_mode)
            
            # CSV输出
            if hasattr(args, 'csv_output') and args.csv_output:
                debtor_name = getattr(args, 'debtor', None)
                append_mode = getattr(args, 'append', False)
                save_csv_output(result, args.csv_output, debtor_name, append_mode)
            
            # JSON输出
            if hasattr(args, 'json_output') and args.json_output:
                output_file = args.json_output if isinstance(args.json_output, str) else None
                save_json_output(result, output_file)
            
            # 如果没有指定输出文件，则表格输出
            if not (hasattr(args, 'excel_output') and args.excel_output) and \
               not (hasattr(args, 'csv_output') and args.csv_output) and \
               not (hasattr(args, 'json_output') and args.json_output):
                print_result_table(result)
            
            # 返回相应的状态码
            sys.exit(result.get('status_code', 0))
        
    except KeyboardInterrupt:
        print("\n程序被用户中断")
        sys.exit(1)
    except Exception as e:
        print(f"❌ 程序运行错误: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main() 